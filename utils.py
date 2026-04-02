import os
import smtplib
from datetime import datetime
from email.mime.text import MIMEText

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def ensure_data_folder(file_path: str) -> None:
    folder = os.path.dirname(file_path)
    if folder and not os.path.exists(folder):
        os.makedirs(folder, exist_ok=True)


def save_price_record(csv_file: str, name: str, price: float, currency: str, url: str) -> None:
    ensure_data_folder(csv_file)

    row = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "product_name": name,
        "price": price,
        "currency": currency,
        "url": url,
    }

    df_new = pd.DataFrame([row])

    if os.path.exists(csv_file) and os.path.getsize(csv_file) > 0:
        df_old = pd.read_csv(csv_file)
        df_all = pd.concat([df_old, df_new], ignore_index=True)
    else:
        df_all = df_new

    df_all.to_csv(csv_file, index=False, encoding="utf-8-sig")


def get_previous_price(csv_file: str, url: str):
    if not os.path.exists(csv_file) or os.path.getsize(csv_file) == 0:
        return None

    df = pd.read_csv(csv_file)
    df = df[df["url"] == url]

    if df.empty:
        return None

    return float(df.iloc[-1]["price"])


def generate_summary(csv_file: str):
    if not os.path.exists(csv_file) or os.path.getsize(csv_file) == 0:
        return []

    df = pd.read_csv(csv_file)

    if df.empty:
        return []

    summary = []

    for url in df["url"].unique():
        product_rows = df[df["url"] == url].copy()

        if product_rows.empty:
            continue

        product_rows = product_rows.reset_index(drop=True)
        latest = product_rows.iloc[-1]

        previous_price = None
        if len(product_rows) >= 2:
            previous_price = float(product_rows.iloc[-2]["price"])

        current_price = float(latest["price"])
        name = str(latest["product_name"])
        currency = str(latest["currency"])
        record_count = len(product_rows)

        if previous_price is None:
            status = "İlk kayıt"
        elif current_price < previous_price:
            status = "Düştü"
        elif current_price > previous_price:
            status = "Arttı"
        else:
            status = "Aynı"

        summary.append({
            "product_name": name,
            "current_price": current_price,
            "previous_price": previous_price,
            "currency": currency,
            "status": status,
            "record_count": record_count,
            "url": url,
        })

    return summary


def style_worksheet(worksheet):
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for column_cells in worksheet.columns:
        max_length = 0
        column = column_cells[0].column
        column_letter = get_column_letter(column)

        for cell in column_cells:
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except Exception:
                pass

        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def color_status_cells(worksheet, status_column_name="status"):
    status_col_index = None

    for idx, cell in enumerate(worksheet[1], start=1):
        if str(cell.value).strip().lower() == status_column_name.lower():
            status_col_index = idx
            break

    if status_col_index is None:
        return

    green_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    red_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFEB9C")
    blue_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")

    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=status_col_index)
        value = str(cell.value).strip().lower()

        if value == "düştü":
            cell.fill = green_fill
        elif value == "arttı":
            cell.fill = red_fill
        elif value == "aynı":
            cell.fill = yellow_fill
        elif value == "ilk kayıt":
            cell.fill = blue_fill


def export_excel_report(csv_file: str, excel_file: str) -> bool:
    if not os.path.exists(csv_file) or os.path.getsize(csv_file) == 0:
        return False

    ensure_data_folder(excel_file)

    history_df = pd.read_csv(csv_file)
    summary_data = generate_summary(csv_file)

    if history_df.empty:
        return False

    summary_df = pd.DataFrame(summary_data)

    with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
        history_df.to_excel(writer, sheet_name="History", index=False)

        if not summary_df.empty:
            summary_df.to_excel(writer, sheet_name="Summary", index=False)

        workbook = writer.book

        history_ws = workbook["History"]
        style_worksheet(history_ws)

        if "Summary" in workbook.sheetnames:
            summary_ws = workbook["Summary"]
            style_worksheet(summary_ws)
            color_status_cells(summary_ws)

    return True


def send_email(subject: str, body: str, config: dict):
    try:
        msg = MIMEText(body, _charset="utf-8")
        msg["Subject"] = subject
        msg["From"] = config["EMAIL_ADDRESS"]
        msg["To"] = config["TO_EMAIL"]

        server = smtplib.SMTP(config["SMTP_SERVER"], config["SMTP_PORT"])
        server.starttls()
        server.login(config["EMAIL_ADDRESS"], config["EMAIL_PASSWORD"])
        server.send_message(msg)
        server.quit()

        print("📧 Mail gönderildi.")
    except Exception as e:
        print(f"Mail hatası: {e}")
        raise